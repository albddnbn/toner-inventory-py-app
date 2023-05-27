# import xlsxwriter
from PyQt5.QtWidgets import (
    QLabel,
    QPushButton,
    QDialog,
    QVBoxLayout,
    QHBoxLayout,
    QSizePolicy,
    QCheckBox,
    QFileDialog,
    QFrame
)
from PyQt5.QtGui import QIcon, QCursor
from PyQt5.QtCore import QSize, Qt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils.toner_utils import connect_to_db, write_log, initialize_config, commit_to_db


class InventoryReport:
    # TODO: go over this entire class, shorten it as much as possible and make sure everything is working as intended
    def __init__(self, report_filename, dell_dict, hp_dict, lexmark_dict, others_dict, col_str: str):
        """
        __init__ Generates an .xlsx spreadsheet report that contains a worksheet for each of the 4 major manufacturer groups of the app ("dell", "hp", "lexmark", and "other").

        Each worksheet contains a table of the devices that belong to that brand.

        Args:
            report_filename (str): path to the report file that will be generated
            dell_dict (dict): dictionary containing the data for the Dell worksheet
            hp_dict (dict): dictionary containing the data for the HP worksheet
            lexmark_dict (dict): dictionary containing the data for the Lexmark worksheet
            others_dict (dict): dictionary containing the data for the Others worksheet
        """
        self.column_string = col_str
        self.columns = self.column_string.split(",")
        self.filename = report_filename
        self.results = {
            "dell": dell_dict,
            "hp": hp_dict,
            "lexmark": lexmark_dict,
            "others": others_dict
        }
        self.dell_dict = dell_dict
        self.hp_dict = hp_dict
        self.lexmark_dict = lexmark_dict
        self.others_dict = others_dict

        self.brand_list = ["Dell", "HP", "Lexmark", "Others"]

        self.worksheets = {"dell": None, "hp": None,
                           "lexmark": None, "others": None}
        self.write_report()

    def write_report(self) -> None:
        wb = Workbook()
        headers = self.column_string.split(",")
        header_font = Font(color='FFFFFF', bold=True, size=14)
        header_fill = PatternFill(
            start_color='3366CC', end_color='3366CC', fill_type='solid')
        self.worksheets["dell"] = wb.active
        self.worksheets["dell"].title = "Dell"

        # cycle through the Excel worksheets, and apply header formatting, column titles
        for key, value in self.worksheets.items():
            if key == "dell":
                self.worksheets["dell"] = wb.active
            else:
                self.worksheets[key] = wb.create_sheet(key.capitalize())
            self.worksheets[key].title = key.capitalize()

            for colnum, headertitle in enumerate(headers, 1):
                column_letter = get_column_letter(colnum)
                self.worksheets[key]['{}1'.format(
                    column_letter)].value = headertitle
                self.worksheets[key]['{}1'.format(
                    column_letter)].font = header_font
                self.worksheets[key]['{}1'.format(
                    column_letter)].fill = header_fill
                self.worksheets[key].column_dimensions[column_letter].width = 20

            for row, result in enumerate(self.results[key], start=1):
                for col, value in result.items():
                    col = self.columns.index(col)
                    self.worksheets[key].cell(
                        row=int(row)+1, column=col+1).value = value


        # for col_num, header_title in enumerate(headers, 1):
        #     column_letter = get_column_letter(col_num)
        #     self.worksheets["dell"]['{}1'.format(
        #         column_letter)].value = header_title
        #     self.worksheets["dell"]['{}1'.format(
        #         column_letter)].font = header_font
        #     self.worksheets["dell"]['{}1'.format(
        #         column_letter)].fill = header_fill
        #     self.worksheets["dell"].column_dimensions[column_letter].width = 20
        # for row, result in enumerate(self.results["dell"], start=1):
        #     for col, value in result.items():
        #         col = self.columns.index(col)
        #         self.worksheets["dell"].cell(
        #             row=int(row)+1, column=col+1).value = value

        # self.worksheets["hp"] = wb.create_sheet("HP")
        # self.worksheets["hp"].title = "HP"
        # for col_num, header_title in enumerate(headers, 1):
        #     column_letter = get_column_letter(col_num)
        #     self.worksheets["hp"]['{}1'.format(
        #         column_letter)].value = header_title
        #     self.worksheets["hp"]['{}1'.format(
        #         column_letter)].font = header_font
        #     self.worksheets["hp"]['{}1'.format(
        #         column_letter)].fill = header_fill
        #     self.worksheets["hp"].column_dimensions[column_letter].width = 20
        # for row, result in enumerate(self.results["hp"], start=1):
        #     # row_format = even_row_format if row % 2 == 0 else odd_row_format
        #     for col, value in result.items():
        #         col = self.columns.index(col)
        #         self.worksheets["hp"].cell(
        #             row=int(row)+1, column=col+1).value = value

        # self.worksheets["lexmark"] = wb.create_sheet("Lexmark")
        # self.worksheets["lexmark"].title = "Lexmark"
        # for col_num, header_title in enumerate(headers, 1):
        #     column_letter = get_column_letter(col_num)
        #     self.worksheets["lexmark"]['{}1'.format(
        #         column_letter)].value = header_title
        #     self.worksheets["lexmark"]['{}1'.format(
        #         column_letter)].font = header_font
        #     self.worksheets["lexmark"]['{}1'.format(
        #         column_letter)].fill = header_fill
        #     self.worksheets["lexmark"].column_dimensions[column_letter].width = 20

        # for row, result in enumerate(self.results["lexmark"], start=1):
        #     for col, value in result.items():
        #         col = self.columns.index(col)
        #         self.worksheets["lexmark"].cell(
        #             row=int(row)+1, column=col+1).value = value

        # self.worksheets["others"] = wb.create_sheet("Others")
        # self.worksheets["others"].title = "Others"
        # for col_num, header_title in enumerate(headers, 1):
        #     column_letter = get_column_letter(col_num)
        #     self.worksheets["others"]['{}1'.format(
        #         column_letter)].value = header_title
        #     self.worksheets["others"]['{}1'.format(
        #         column_letter)].font = header_font
        #     self.worksheets["others"]['{}1'.format(
        #         column_letter)].fill = header_fill
        #     self.worksheets["others"].column_dimensions[column_letter].width = 20
        # for row, result in enumerate(self.results["others"], start=1):
        #     for col, value in result.items():
        #         col = self.columns.index(col)
        #         self.worksheets["others"].cell(
        #             row=int(row)+1, column=col+1).value = value
        # iterate through rows and apply alternating row colors
        border = Border(left=Side(style='medium', color='FFFFFF'),
                        right=Side(style='medium', color='FFFFFF'),
                        top=Side(style='medium', color='FFFFFF'),
                        bottom=Side(style='medium', color='FFFFFF'))
        # apply alternating row colors
        for worksheet in self.worksheets.values():
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.row % 2:
                        # cell.fill = PatternFill(start_color='7392ff', end_color='7392ff', fill_type="solid")
                        # set font to white color
                        cell.font = Font(
                            name='Roboto', color='000000', size=12)
                    else:
                        cell.fill = PatternFill(
                            start_color='DEEBFE', end_color='DEEBFE', fill_type="solid")
                        cell.font = Font(
                            name='Roboto', color='000000', size=12)
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=True)
                    cell.border = border
        # set column widths
        for worksheet in self.worksheets.values():
            col_num = 1
            while worksheet.cell(row=1, column=col_num).value:
                col_letter = get_column_letter(col_num)

                # get text value of header
                # header_text = worksheet.cell(row=1, column=col_num).value
                header_text = worksheet[f"{col_letter}1"].value

                # if header_text is description, set column width to 50
                if header_text.lower() == "description":
                    worksheet.column_dimensions[col_letter].width = 50
                    worksheet.column_dimensions[col_letter].alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=True)
                elif header_text.lower() == "printers":
                    worksheet.column_dimensions[col_letter].width = 40
                elif header_text.lower() in ["stock", "needed"]:
                    worksheet.column_dimensions[col_letter].width = 8
                elif header_text.lower() == "code":
                    worksheet.column_dimensions[col_letter].width = 20
                elif header_text.lower() == "brand":
                    worksheet.column_dimensions[col_letter].width = 10

                col_num += 1

        # close the workbook
        wb.save(self.filename)


class ReportDialog(QDialog):
    def __init__(self, parent=None):
        """
        __init__ Meant to display a popup to the user which asks them to choose which type of report they'd like generated.

        Args:
            results (list, optional): _description_. Defaults to [].
            parent (_type_, optional): _description_. Defaults to None.
        """
        super(ReportDialog, self).__init__(parent=None)
        initialize_config(self)
        self.setWindowTitle('Choose columns & report type')
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.setWindowIcon(QIcon(self.icon))
        self.columns = {
            "0": ["brand", "mfr"],
            "1": ["description", "item name"],
            "2": ["printers", "compatible printers"],
            "3": ["code"],
            "4": ["upc"],
            "5": ["stock"],
            "6": ["part_num", "part number"],
            "7": ["color"],
            "8": ["yield"],
            "9": ["type", "t"],
            "10": ["comments"],
            "11": ["linked_codes", "linked items"],
            "12": ["needed"]
        }

        # create the layout and widgets
        layout = QVBoxLayout()
        label = QLabel('Choose columns and type of report:')
        # hbox for color choices

        label.setProperty("texttype", "title")
        layout.addWidget(label)
        layout.setAlignment(label, Qt.AlignCenter)

        checkbox_hbox = QHBoxLayout()
        vframe1 = QFrame()
        vbox1 = QVBoxLayout()
        vframe1.setLayout(vbox1)
        vframe1.setProperty("class", "col-frame")
        vframe2 = QFrame()
        vbox2 = QVBoxLayout()
        vframe2.setLayout(vbox2)
        vframe2.setProperty("class", "col-frame")
        # iterator to make sure half of the checkboxes are in the first vbox
        x = 1
        # create a list of checkboxes
        checkboxes = {}
        for name_list in self.columns.values():
            column_name = name_list[0]
            checkboxes[column_name] = QCheckBox(column_name)
            if x <= len(self.columns) / 2:
                vbox1.addWidget(checkboxes[column_name])
            else:
                vbox2.addWidget(checkboxes[column_name])
            x += 1
            if name_list[0] == "stock":
                checkboxes["stock"].setChecked(True)
        # checkbox_hbox.addLayout(vbox1)
        # checkbox_hbox.addLayout(vbox2)
        checkbox_hbox.addWidget(vframe1)
        checkbox_hbox.addWidget(vframe2)
        layout.addLayout(checkbox_hbox)
        layout.setAlignment(checkbox_hbox, Qt.AlignCenter)

        button_layout = QHBoxLayout()

        self.bybrandbtn = QPushButton('By Manufacturer')
        self.bybrandbtn.setCursor(QCursor(Qt.PointingHandCursor))
        self.bybrandbtn.setToolTip(
            'Items arranged in worksheets by manufacturer')

        # low items report button
        low_items_btn = QPushButton('Low stock items')
        # low_items_btn.setProperty("class", "report-btn")

        low_items_btn.setCursor(QCursor(Qt.PointingHandCursor))
        low_items_btn.setToolTip(
            'Items with stock of 2 or less arranged in worksheets by manufacturer')

        needed_btn = QPushButton('Items needed')
        needed_btn.setCursor(QCursor(Qt.PointingHandCursor))
        needed_btn.setToolTip(
            'If column has >0 needed value, and stock is < needed, item included in report')

        button_layout.addWidget(self.bybrandbtn)
        button_layout.addWidget(low_items_btn)
        button_layout.addWidget(needed_btn)
        # add the widgets to the layout
        layout.addLayout(button_layout)
        # set the layout for the dialog
        self.setLayout(layout)
        # connect buttons to generate report functions
        self.bybrandbtn.clicked.connect(
            lambda: self.return_choices(checkboxes=checkboxes))
        self.bybrandbtn.clicked.connect(self.positive_exit)
        low_items_btn.clicked.connect(lambda: self.return_choices(
            checkboxes=checkboxes, report_type="lowitems"))
        low_items_btn.clicked.connect(self.positive_exit)

        needed_btn.clicked.connect(lambda: self.return_choices(
            checkboxes=checkboxes, report_type="needed"))
        needed_btn.clicked.connect(self.positive_exit)

        button_layout.setSpacing(10)
        button_layout.setAlignment(Qt.AlignCenter)
        button_layout.setContentsMargins(15, 5, 15, 5)
        # self.setStyleSheet(parent.styleSheet())

    def positive_exit(self):

        self.accept()

    def sizeHint(self):
        return QSize(500, 150)
    # generate a report by manufacturer

    def return_choices(self, checkboxes: dict, report_type="bybrand"):
        checked_columns = []
        column_string = ""
        for column_name in checkboxes.keys():
            if checkboxes[column_name].isChecked():
                column_string += f"{column_name},"
                checked_columns.append(column_name)
        if column_string.endswith(","):
            column_string = column_string[:-1]

        self.column_string = column_string
        reportfile, _ = QFileDialog.getSaveFileName(
            None, "Save File", "", "Excel files (*.xlsx)"
        )

        if (reportfile != "") and (reportfile != None):
            print("test")
            # try:
            database, cursorObject = connect_to_db(
                db_host=self.dbhost,
                db_user=self.dbuser,
                db_name=self.dbname,
                db_pass=self.dbpass,
                dicttrue=True,
            )
            queries = {}
            for brandname in ["dell", "hp", "lexmark"]:
                if report_type == "bybrand":
                    queries[brandname] = [
                        f"SELECT {column_string} FROM inventory WHERE LOWER(brand)='{brandname}' ORDER BY brand, stock DESC, printers"
                    ]
                elif report_type == "lowitems":
                    queries[brandname] = [
                        f"SELECT {column_string} FROM inventory WHERE LOWER(brand)='{brandname}' AND stock <= 2 ORDER BY stock DESC"
                    ]
                elif report_type == "needed":
                    queries[brandname] = [
                        f"SELECT {column_string} FROM inventory WHERE (LOWER(brand)='{brandname}') AND (needed > 0) AND (stock < needed) ORDER BY stock DESC"
                    ]
                cursorObject.execute(queries[brandname][0])
                # append the results of the query to 1st index of the list
                queries[brandname].append(cursorObject.fetchall())

            dell_results = queries["dell"][1]
            hp_results = queries["hp"][1]
            lexmark_results = queries["lexmark"][1]

            # get the items that aren't dell, hp, or lexmark
            if (report_type == "bybrand") and ("stock" in checked_columns):
                others_query = f"SELECT {column_string} FROM inventory WHERE LOWER(brand) NOT IN ('dell', 'hp', 'lexmark') ORDER BY brand, printers"
            elif (report_type == "lowitems") and ("stock" in checked_columns):
                print("setting others query")
                others_query = f"SELECT {column_string} FROM inventory WHERE (LOWER(brand) NOT IN ('dell', 'hp', 'lexmark')) AND (stock <= 2) ORDER BY stock DESC"
            elif (report_type == "needed") and ("stock" in checked_columns):
                others_query = f"SELECT {column_string} FROM inventory WHERE (LOWER(brand) NOT IN ('dell', 'hp', 'lexmark')) AND (needed > 0) AND (stock < needed) ORDER BY stock DESC"
            cursorObject.execute(others_query)
            others_results = cursorObject.fetchall()
            InventoryReport(reportfile, dell_results,
                            hp_results, lexmark_results, others_results, column_string)
            # except Exception as e:
            #     print(e)

    def gen_bybrand_report(self, checkboxes: dict):
        """
        gen_bybrand_report generates report of printer-related inventory in excel worksheets by manufacturer (dell, hp, lexmark, others) by calling the InventoryReport class
        """
        # get the text from the checked textboxes
        checked_columns = []
        column_string = ""
        for column_name in checkboxes.keys():
            if checkboxes[column_name].isChecked():
                column_string += f"{column_name},"
                checked_columns.append(column_name)
        if column_string.endswith(","):
            column_string = column_string[:-1]
        InventoryReport(self.filename, self.dell_results,
                        self.hp_results, self.lexmark_results, self.others_results, column_string)

    def gen_low_items_report(self, checkboxes: dict):
        '''
        gen_low_items_report generate a report of items with stock of 2 or less.

        Like the basic report, this report is arranged in excel worksheets by manufacturer.
        '''
        checked_columns = []
        column_string = ""
        for column_name in checkboxes.keys():
            if checkboxes[column_name].isChecked():
                column_string += f"{column_name},"
                checked_columns.append(column_name)
        if column_string.endswith(","):
            column_string = column_string[:-1]
        InventoryReport(self.filename, self.dell_results,
                        self.hp_results, self.lexmark_results, self.others_results, column_string)
    
    def gen_needed_report(self, checkboxes: dict):
        '''
        gen_needed_report generates a report of items that need to be ordered.

        This report is arranged in excel worksheets by manufacturer.
        '''
        checked_columns = []
        column_string = ""
        for column_name in checkboxes.keys():
            if checkboxes[column_name].isChecked():
                column_string += f"{column_name},"
                checked_columns.append(column_name)
        if column_string.endswith(","):
            column_string = column_string[:-1]
        InventoryReport(self.filename, self.dell_results,
                        self.hp_results, self.lexmark_results, self.others_results, column_string)